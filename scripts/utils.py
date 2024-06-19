# Importing useful libraries
import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas as pd
import numpy as np
from io import BytesIO


# Function to apply formatting to a sheet
def format_sheet(sheet):
    # Color the first row black
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_bold_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = black_fill
        cell.font = white_bold_font # Change font color to white for readability

    # Coloring the cell containing the total data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Total':
                cell.fill = black_fill
                cell.font = white_bold_font

    # Adjust column widths to fit the contents
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

def generate_report(input_data_path,save_path):

    # Reading the revenue data
    xls = pd.read_excel(input_data_path, sheet_name=None)
    data_frames = {}

    # xls is a dictionary where the keys are the sheet names and the values are the DataFrames
    for sheet_name, sheet_df in xls.items():
        data_frames[sheet_name] = sheet_df

    list_of_campaigns = []

    for key in data_frames.keys():

        cut_off_1 = data_frames[key][data_frames[key]['Campaign'] == 'Publisher Name'].index[0]
        temp_df = data_frames[key].iloc[cut_off_1:]
        cut_off_2 = temp_df[temp_df['Campaign'].isna() == True].index[0]
        campaigns = list(set(list(data_frames[key].iloc[cut_off_1:cut_off_2]['Campaign'].values)))
        campaigns.remove('Publisher Name')

        list_of_campaigns += campaigns

    campaigns = list(set(list_of_campaigns))

    report_data = {

    }

    for campaign in campaigns:

        temp_campaign_df = {

        }

        for key in data_frames.keys():

            temp_df = data_frames[key][data_frames[key]['Campaign'] == campaign]
            useful_columns = ['Publisher', 'Campaign', 'Leads', 'Revenue', 'Unnamed: 4', 'Unnamed: 5','Unnamed: 6']
            temp_df = temp_df[useful_columns]
            temp_df['Date'] = [key] * temp_df.shape[0]
            temp_df.rename(
                columns = {
                    'Unnamed: 4' : 'Clicks/Views', 
                    'Unnamed: 5' : 'We Pay',
                    'Unnamed: 6' : 'Margin'
                },
                inplace=True
            )

            temp_df = temp_df[
                ['Date',
                'Publisher',
                'Campaign',
                'Leads',
                'Revenue',
                'Clicks/Views',
                'We Pay',
                'Margin']
            ]

            # Convert the 'Margin' column to percentage values
            temp_df['Margin'] = temp_df['Margin'].apply(lambda x: f'{int(x*100)}%' if pd.notnull(x) else '0.0%')

            temp_campaign_df[key] = temp_df

        temp_dfs = []

        for key,value in temp_campaign_df.items():

            temp_df = value.copy()

            sub_ids = [id for id in list(set(temp_campaign_df[key]['Publisher'].values)) if '/' in str(id)]

            grouped_dfs = []
            
            try:

                for id in list(set(sub_ids)):

                    grouped_dfs.append(temp_df[temp_df['Publisher'] == id])

            

                temp_df = pd.concat(grouped_dfs,axis = 0)

            except:

                pass

            temp_dfs.append(value)

        blank_row = pd.DataFrame([[np.nan] * temp_df.shape[1]], columns=temp_df.columns)

        # Initialize an empty list to hold the DataFrames with blank rows in between
        concatenated_dfs = []

        # Loop through the DataFrames and concatenate with a blank row in between
        for df in temp_dfs:
            concatenated_dfs.append(df)
            concatenated_dfs.append(blank_row)

        temp_final_df = pd.concat(concatenated_dfs,axis = 0)    

        report_data[campaign] = temp_final_df

    final_report_df = {}

    for key in report_data.keys():

        grouped_dfs = []

        sub_ids = [id for id in list(set(report_data[key]['Publisher'].values)) if '/' in str(id)]

        temp_df = report_data[key]

        for id in list(set(sub_ids)):

            grouped_dfs.append(temp_df[temp_df['Publisher'] == id])

            total_leads = temp_df[temp_df['Publisher'] == id]['Leads'].sum()
            total_revenue = temp_df[temp_df['Publisher'] == id]['Revenue'].sum()
            total_clicks = temp_df[temp_df['Publisher'] == id]['Clicks/Views'].sum()
            total_cost = temp_df[temp_df['Publisher'] == id]['We Pay'].sum()

            try:
                total_margin = f'{round(((total_revenue - total_cost) / total_revenue)*100,0)}%'
            except:
                total_margin = '0.0%'

            total_data = ['','','Total',total_leads,total_revenue,total_clicks,total_cost,total_margin]

            total_data_df = pd.DataFrame([total_data],
                                         columns=temp_df.columns)

            grouped_dfs.append(total_data_df) 

            grouped_dfs.append(blank_row)

        try: temp_df = pd.concat(grouped_dfs,axis = 0)
        except: pass

        final_report_df[key] = temp_df

    # Step 2: Sort the keys
    sorted_keys = sorted(final_report_df.keys())

    # Step 3: Create a new dictionary with sorted keys
    sorted_dict = {key: final_report_df[key] for key in sorted_keys}

    excel_path = save_path

    final_report_df = sorted_dict

    # Create an Excel file with multiple sheets
    with pd.ExcelWriter(excel_path) as writer:

        for name,df in final_report_df.items():

            df.to_excel(writer, sheet_name=name, index=False)

    # Save the DataFrame to an Excel file
    excel_path = excel_path

    # Load the Excel file with openpyxl
    workbook = openpyxl.load_workbook(excel_path)

    # Apply formatting to all sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        format_sheet(sheet)

    workbook.save(excel_path)

def generate_excel_file(dataframes):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            # Access the workbook and worksheet
            workbook = writer.book
            worksheet = workbook[sheet_name]
            # Apply styles
            format_sheet(worksheet)
    output.seek(0)  # Move the cursor back to the start of the BytesIO object
    return output

